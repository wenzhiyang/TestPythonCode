import datetime
import calendar
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from SNFunctionList import *
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
from openpyxl.utils import range_boundaries
from openpyxl.formula.translate import Translator

import os
import pandas as pd
import mysql.connector

from datetime import datetime, timedelta
import re
from openpyxl import Workbook
import calendar
from dateutil.relativedelta import relativedelta
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
# 连接到 MySQL 数据库
from pycel import ExcelCompiler

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
import re



common_columns = [
         '2#', '转S1-1#', '2-1#', '2-1T#', '2A-1TB2#', '2A-2#', '2A',
        'R6#', '石S1-1#', '石S1-1T#', 'GHBA-4#', '针S1-1#', '针S1-1T#',
        'S1-2#', 'S1-3T#', '高S2-2#', 'S2-2T#', 'S2-1#', '低S2-2#','S2-5#',
        'S4-1#','S5-1#'
    ]
customer_dict = {
    'A001': '安德丰',
    'A002': '安普瑞斯',
    'A003': '埃克森',
    'B001': '保力新',
    'B002': '比亚迪',
    'G001': '国轩',
    'J001': '力神聚元',
    'L001': '力神',
    'L003': '朗泰通',
    'L004': '朗泰沣',
    'N002': '宁德时代',
    'N003': '能优',
    'S003': '索理德',
    'S004':'三一',
    'T002': '拓邦',
    'T004': '天鹏',
    'T005': '天弋',
    'W001':'沃能',
    'W002': '五行',
    'X001':'鑫辉',
    'X003': '星恒'
}
def parse_rateorder_file(file_path, plan_file_path):
    """读取配比单文件并处理数据，增加计划部文件路径参数"""
    try:
        # 初始化分配详情列表
        allocation_details = []

        # 读取计划部文件获取计划数
        plan_df = pd.read_excel(
            plan_file_path,
            sheet_name='2025年6月成品排产跟踪',
            skiprows=2,  # 跳过前两行（标题和空行）
            usecols='A:B,D',  # 客户名称(A), 产品型号(B), 计划数(D)
            names=['客户名称', '产品型号', '计划数']
        )
        plan_df['计划数'] = pd.to_numeric(plan_df['计划数'], errors='coerce').fillna(0)

        # 修改：增加配方编号列(A列)
        df = pd.read_excel(
            file_path,
            sheet_name='每日配比单进度跟踪',
            skiprows=1,
            usecols='B,D:H',  # 增加A列(配方编号)
            header=None,
            names=['配方编号', '产品型号', '客户代码', '配比单生产总量(公斤)', '已生产数量', '备注']
        )

        # 筛选备注不包含"生产完"的行
        df = df[df['备注'].fillna('').str.contains('生产完') == False]

        # 计算剩余生产量
        df['剩余生产量'] = (df['配比单生产总量(公斤)'] - df['已生产数量']) / 1000
        df.reset_index(drop=True, inplace=True)

        # 拆分客户代码列
        df_expanded_list = []

        for idx, row in df.iterrows():
            formula_code = row['配方编号']  # 获取配方编号
            customer_codes = str(row['客户代码']).replace(',', ' ').split()
            pattern = r'^[A-Za-z]\d{3}$'
            valid_codes = [code for code in customer_codes if pd.Series(code).str.match(pattern)[0]]

            if not valid_codes:
                continue

            # 单个客户代码处理
            if len(valid_codes) == 1:
                new_row = row.copy()
                new_row['客户代码'] = valid_codes[0]
                new_row['客户名称'] = customer_dict.get(valid_codes[0], '')
                df_expanded_list.append(new_row)
                continue

            # 多个客户代码需要分配剩余生产量
            total_remaining = row['剩余生产量']
            customer_data = []

            # 获取每个客户的计划数
            for code in valid_codes:
                customer_name = customer_dict.get(code, '')
                plan_value = 0
                if customer_name and row['产品型号']:
                    match = plan_df[
                        (plan_df['客户名称'] == customer_name) &
                        (plan_df['产品型号'] == row['产品型号'])
                        ]
                    if not match.empty:
                        plan_value = match['计划数'].iloc[0]
                customer_data.append({
                    '客户代码': code,
                    '客户名称': customer_name,
                    '计划数': plan_value
                })

            # 新分配逻辑：按顺序分配给计划数>0的客户，最后未分配完的给最后一个客户
            allocations = [0] * len(customer_data)  # 初始化分配列表
            allocated = 0  # 已分配量

            # 按顺序遍历客户
            for i, cust in enumerate(customer_data):
                # 如果还有剩余量且当前客户有计划数
                if allocated < total_remaining and cust['计划数'] > 0:
                    # 可分配量 = min(客户计划数, 剩余总量 - 已分配量)
                    alloc_amount = min(cust['计划数'], total_remaining - allocated)
                    allocations[i] = alloc_amount
                    allocated += alloc_amount

            # 如果还有剩余未分配，分配给最后一个客户
            if allocated < total_remaining:
                # 计算剩余量
                remaining = total_remaining - allocated
                # 分配给最后一个客户
                allocations[-1] += remaining
                allocated += remaining

            # 记录分配详情（仅多客户情况）
            for i, cust in enumerate(customer_data):
                allocation_details.append({
                    '配方编号': formula_code,
                    '产品型号': row['产品型号'],
                    '客户代码': cust['客户代码'],
                    '客户名称': cust['客户名称'],
                    '分配数量': allocations[i]
                })

            # 创建拆分后的行
            for i, cust in enumerate(customer_data):
                new_row = row.copy()
                new_row['客户代码'] = cust['客户代码']
                new_row['客户名称'] = cust['客户名称']
                new_row['剩余生产量'] = allocations[i]
                df_expanded_list.append(new_row)

        # 合并所有行
        if df_expanded_list:
            df_expanded = pd.DataFrame(df_expanded_list)
        else:
            df_expanded = pd.DataFrame(columns=df.columns.tolist() + ['剩余生产量', '客户名称'])

        # 分组汇总剩余生产量
        grouped = df_expanded.groupby(['客户代码', '客户名称', '产品型号'], as_index=False).agg(
            totalproding=('剩余生产量', 'sum')
        )

        # 将分配详情转换为DataFrame
        allocation_df = pd.DataFrame(allocation_details)

        return grouped, allocation_df

    except Exception as e:
        print(f"处理文件出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(columns=['客户代码', '客户名称', '产品型号', 'totalproding']), pd.DataFrame(columns=['配方编号', '产品型号', '客户代码', '客户名称', '分配数量'])

def update_production_trace(sales_df, production_file_path, sheetname,nextmonthsheetname, nm_prodingfile, hb_prodingfile,
                            NM_data_storage, HB_data_storage, nm_data_prodsend, hb_data_prodsend,
                            NM_all_storage, HB_all_storage,NM_return_data,HB_return_data):
    """
    更新生产跟踪文件并生成新版本（支持多列更新）
    :param sales_df: 销售数据DataFrame（需包含客户、产品型号、当月列、内蒙计划数、湖北计划数）
    :return: 新生成的文件路径
    """
    # 生成新文件名
    new_file_path = generate_new_filename(production_file_path)
    smart_decimal_format = '#,##0.###'

    try:
        # 复制模板文件
        shutil.copyfile(production_file_path, new_file_path)

        # 加载新文件的工作表
        wb = load_workbook(filename=new_file_path)
        if sheetname not in wb.sheetnames:
            raise ValueError(f"工作表 {sheetname} 不存在")
        ws = wb[sheetname]

        # 2. 处理下月工作表
        if nextmonthsheetname not in wb.sheetnames:
            raise ValueError(f"工作表 {nextmonthsheetname} 不存在")
        ws_next = wb[nextmonthsheetname]

        # 处理内蒙配比单文件
        nm_proding_data, nm_allocation_details = parse_rateorder_file(nm_prodingfile, new_file_path)

        # 处理湖北配比单文件
        hb_proding_data, hb_allocation_details = parse_rateorder_file(hb_prodingfile, new_file_path)

        # 提取当月列名和下月列名
        current_month_col = sales_df.columns[2]  # 当月销售计划列名
        next_month_col = sales_df.columns[5]  # 下月销售计划列名
        next_month_nm_col = sales_df.columns[6]  # 下月内蒙计划数列名
        next_month_hb_col = sales_df.columns[7]  # 下月湖北计划数列名

        # 构建销售数据索引
        # sales_dict = sales_df.set_index(['客户', '产品型号'])[
        #     [sales_df.columns[2], '内蒙计划数', '湖北计划数']].to_dict(orient='index')
        sales_dict = sales_df.set_index(['客户', '产品型号'])[
            [current_month_col, '内蒙计划数', '湖北计划数',
             next_month_col, next_month_nm_col, next_month_hb_col]
        ].to_dict(orient='index')

        # 构建配比单数据索引
        nm_proding_dict = nm_proding_data.set_index(['客户名称', '产品型号'])['totalproding'].to_dict()
        hb_proding_dict = hb_proding_data.set_index(['客户名称', '产品型号'])['totalproding'].to_dict()

        # 初始化配比单未匹配集合
        nm_unmatched_rate = set(nm_proding_dict.keys())
        hb_unmatched_rate = set(hb_proding_dict.keys())

        # 初始化其他未匹配数据容器
        nm_unmatched = NM_data_storage[['客户', '产品型号', '库存']].to_dict('records')
        hb_unmatched = HB_data_storage[['客户', '产品型号', '库存']].to_dict('records')
        nm_send_unmatched = nm_data_prodsend[['客户', '产品型号', '内蒙发货数']].to_dict('records')
        hb_send_unmatched = hb_data_prodsend[['客户', '产品型号', '湖北发货数']].to_dict('records')

        # 转换为集合以便快速删除
        nm_unmatched_set = {(item['客户'], item['产品型号']) for item in nm_unmatched}
        hb_unmatched_set = {(item['客户'], item['产品型号']) for item in hb_unmatched}
        nm_send_unmatched_set = {(item['客户'], item['产品型号']) for item in nm_send_unmatched}
        hb_send_unmatched_set = {(item['客户'], item['产品型号']) for item in hb_send_unmatched}

        # 构建索引字典
        nm_stock = NM_data_storage.drop_duplicates(subset=['客户', '产品型号'], keep='last')
        nm_stock_dict = nm_stock.set_index(['客户', '产品型号']).to_dict('index')

        hb_stock = HB_data_storage.drop_duplicates(subset=['客户', '产品型号'], keep='last')
        hb_stock_dict = hb_stock.set_index(['客户', '产品型号']).to_dict('index')

        nm_send = nm_data_prodsend.drop_duplicates(subset=['客户', '产品型号'], keep='last')
        nm_send_dict = nm_send.set_index(['客户', '产品型号']).to_dict('index')

        hb_send = hb_data_prodsend.drop_duplicates(subset=['客户', '产品型号'], keep='last')
        hb_send_dict = hb_send.set_index(['客户', '产品型号']).to_dict('index')

        # 遍历跟踪数据 处理当月工作表
        updated_rows = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
            customer = row[0].value  # A列 - 客户名称
            product = row[1].value  # B列 - 产品型号

            # 终止条件：A/B列同时为空
            if customer is None and product is None:
                break

            key = (customer, product)

            # 匹配销售数据 - 列索引调整（新增包装方式列）
            if key in sales_dict:
                # 获取三列数据
                month_value = sales_dict[key][current_month_col]
                nm_plan = sales_dict[key]['内蒙计划数']
                hb_plan = sales_dict[key]['湖北计划数']

                # 更新三列（D=4, E=5, F=6）
                # D列：6月销售计划
                ws.cell(row=row_idx, column=4, value=month_value)

                # E列：内蒙计划数
                ws.cell(row=row_idx, column=5, value=nm_plan).number_format = '0'

                # F列：湖北计划数
                ws.cell(row=row_idx, column=6, value=hb_plan).number_format = '0'

                # 设置格式
                for col in [4, 5, 6]:
                    ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_idx, column=col).font = Font(bold=True, size=10)

                updated_rows += 1

            # 处理内蒙发货 - 列索引调整（H=8）
            if key in nm_send_dict:
                ws.cell(row=row_idx, column=8,
                        value=nm_send_dict[key]['内蒙发货数']).number_format = smart_decimal_format
                ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=8).font = Font(bold=True, size=10)
                if key in nm_send_unmatched_set:
                    nm_send_unmatched_set.remove(key)

            # 处理湖北发货 - 列索引调整（I=9）
            if key in hb_send_dict:
                ws.cell(row=row_idx, column=9,
                        value=hb_send_dict[key]['湖北发货数']).number_format = smart_decimal_format
                ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=9).font = Font(bold=True, size=10)
                if key in hb_send_unmatched_set:
                    hb_send_unmatched_set.remove(key)

            # 处理内蒙库存 - 列索引调整（K=11）
            if key in nm_stock_dict:
                ws.cell(row=row_idx, column=11, value=nm_stock_dict[key]['库存']).number_format = smart_decimal_format
                ws.cell(row=row_idx, column=11).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=11).font = Font(bold=True, size=10)
                if key in nm_unmatched_set:
                    nm_unmatched_set.remove(key)

            # 处理湖北库存 - 列索引调整（L=12）
            if key in hb_stock_dict:
                ws.cell(row=row_idx, column=12, value=hb_stock_dict[key]['库存']).number_format = smart_decimal_format
                ws.cell(row=row_idx, column=12).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=12).font = Font(bold=True, size=10)
                if key in hb_unmatched_set:
                    hb_unmatched_set.remove(key)

            # 处理配比单生产数据
            rate_key = key
            if product == "SN-LTF":
                rate_key = (customer, "SN-P2C-1")

            # N列(14列) - 内蒙配比单生产汇总
            if rate_key in nm_proding_dict:
                ws.cell(row=row_idx, column=14, value=nm_proding_dict[rate_key]).number_format = smart_decimal_format
                ws.cell(row=row_idx, column=14).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=14).font = Font(bold=True, size=10)
                # 从配比单未匹配集合中移除已匹配的项
                if rate_key in nm_unmatched_rate:
                    nm_unmatched_rate.remove(rate_key)

            # O列(15列) - 湖北配比单生产汇总
            if rate_key in hb_proding_dict:
                ws.cell(row=row_idx, column=15, value=hb_proding_dict[rate_key]).number_format = smart_decimal_format
                ws.cell(row=row_idx, column=15).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row_idx, column=15).font = Font(bold=True, size=10)
                # 从配比单未匹配集合中移除已匹配的项
                if rate_key in hb_unmatched_rate:
                    hb_unmatched_rate.remove(rate_key)

        # 2. 收集当月工作表的数据用于下月库存计算
        prev_sheet_data = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
            customer = row[0].value
            product = row[1].value
            if customer is None and product is None:
                break

            def to_float(value):
                if value is None:
                    return 0.0
                try:
                    return float(value)
                except (TypeError, ValueError):
                    return 0.0

            # # 读取内蒙相关列
            # nm_plan = row[4].value or 0  # E列 - 内蒙计划数
            # nm_sent = row[7].value or 0  # H列 - 内蒙发货数
            # nm_stock = row[10].value or 0  # K列 - 内蒙库存
            # nm_rate = row[13].value or 0  # N列 - 内蒙配比单生产
            # nm_transit = row[16].value or 0  # Q列 - 内蒙在途
            #
            # # 读取湖北相关列
            # hb_plan = row[5].value or 0  # F列 - 湖北计划数
            # hb_sent = row[8].value or 0  # I列 - 湖北发货数
            # hb_stock = row[11].value or 0  # L列 - 湖北库存
            # hb_rate = row[14].value or 0  # O列 - 湖北配比单生产
            # hb_transit = row[15].value or 0  # P列 - 湖北在途

            # 读取内蒙相关列
            nm_plan_val = to_float(row[4].value)  # E列 - 内蒙计划数
            nm_sent_val = to_float(row[7].value)  # H列 - 内蒙发货数
            nm_stock_val = to_float(row[10].value)  # K列 - 内蒙库存
            nm_rate_val = to_float(row[13].value)  # N列 - 内蒙配比单生产
            nm_transit_val = to_float(row[16].value)  # Q列 - 内蒙在途

            # 读取湖北相关列
            hb_plan_val = to_float(row[5].value)  # F列 - 湖北计划数
            hb_sent_val = to_float(row[8].value)  # I列 - 湖北发货数
            hb_stock_val = to_float(row[11].value)  # L列 - 湖北库存
            hb_rate_val = to_float(row[14].value)  # O列 - 湖北配比单生产
            hb_transit_val = to_float(row[15].value)  # P列 - 湖北在途

            prev_sheet_data[(customer, product)] = {
                'nm_plan': nm_plan_val,
                'nm_sent': nm_sent_val,
                'nm_stock': nm_stock_val,
                'nm_rate': nm_rate_val,
                'nm_transit': nm_transit_val,
                'hb_plan': hb_plan_val,
                'hb_sent': hb_sent_val,
                'hb_stock': hb_stock_val,
                'hb_rate': hb_rate_val,
                'hb_transit': hb_transit_val
            }

        # ===== 处理下月工作表 =====
        updated_rows_next = 0
        for row_idx, row in enumerate(ws_next.iter_rows(min_row=4), start=4):
            customer = row[0].value  # A列 - 客户名称
            product = row[1].value  # B列 - 产品型号

            # 终止条件：A/B列同时为空
            if customer is None and product is None:
                break

            key = (customer, product)

            # 匹配销售数据 - 只写入计划数据
            if key in sales_dict:
                # 获取三列数据（下月）
                next_month_value = sales_dict[key][next_month_col]
                next_month_nm = sales_dict[key][next_month_nm_col]
                next_month_hb = sales_dict[key][next_month_hb_col]

                # 更新三列（D=4, E=5, F=6）
                # D列：下月销售计划
                ws_next.cell(row=row_idx, column=4, value=next_month_value)

                # E列：下月内蒙计划数
                ws_next.cell(row=row_idx, column=5, value=next_month_nm).number_format = '0'

                # F列：下月湖北计划数
                ws_next.cell(row=row_idx, column=6, value=next_month_hb).number_format = '0'

                # 设置格式
                for col in [4, 5, 6]:
                    ws_next.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center',
                                                                                vertical='center')
                    ws_next.cell(row=row_idx, column=col).font = Font(bold=True, size=10)

                updated_rows_next += 1

            if key in prev_sheet_data:
                data = prev_sheet_data[key]

                # 计算内蒙库存 (K列)
                nm_total = (data['nm_stock'] +
                            data['nm_rate'] +
                            data['nm_transit'] -
                            data['nm_plan'] +
                            data['nm_sent'])
                nm_total = max(0, nm_total)  # 小于0时取0

                # 计算湖北库存 (L列)
                hb_total = (data['hb_stock'] +
                            data['hb_rate'] +
                            data['hb_transit'] -
                            data['hb_plan'] +
                            data['hb_sent'])
                hb_total = max(0, hb_total)  # 小于0时取0

                # 写入计算结果
                ws_next.cell(row=row_idx, column=11, value=nm_total).number_format = smart_decimal_format
                ws_next.cell(row=row_idx, column=12, value=hb_total).number_format = smart_decimal_format

                # 设置格式
                for col in [11, 12]:
                    cell = ws_next.cell(row=row_idx, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True, size=10)


        # 更新未匹配数据列表
        nm_unmatched = [item for item in nm_unmatched if (item['客户'], item['产品型号']) in nm_unmatched_set]
        hb_unmatched = [item for item in hb_unmatched if (item['客户'], item['产品型号']) in hb_unmatched_set]
        nm_send_unmatched = [item for item in nm_send_unmatched if
                             (item['客户'], item['产品型号']) in nm_send_unmatched_set]
        hb_send_unmatched = [item for item in hb_send_unmatched if
                             (item['客户'], item['产品型号']) in hb_send_unmatched_set]

        # 准备配比单未匹配数据
        nm_rate_unmatched = [
            {'客户': key[0], '产品型号': key[1], '剩余生产量': value}
            for key, value in nm_proding_dict.items()
            if key in nm_unmatched_rate
        ]

        hb_rate_unmatched = [
            {'客户': key[0], '产品型号': key[1], '剩余生产量': value}
            for key, value in hb_proding_dict.items()
            if key in hb_unmatched_rate
        ]

        # 写入未匹配数据到新sheet（带格式设置）
        def write_unmatched_sheet(wb, data, sheet_name, columns):

            if not data:
                print(f"未匹配数据 '{sheet_name}' 为空，跳过创建sheet")
                return

            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append(columns)

            # 确定数值列索引
            numeric_columns = ['库存', '内蒙发货数', '湖北发货数','库存（吨）','分配数量']
            numeric_indices = [i for i, col in enumerate(columns) if col in numeric_columns]

            for item in data:
                row = [item.get(col, 0) for col in columns]
                ws.append(row)

                # 设置数值列格式
                for col_idx in numeric_indices:
                    cell = ws.cell(row=ws.max_row, column=col_idx + 1)  # openpyxl列从1开始
                    cell.number_format = smart_decimal_format

        def write_allocation_sheet(wb, details_df, sheet_name):
            if not details_df.empty:
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                ws = wb.create_sheet(sheet_name)

                # 添加标题
                headers = ['配方编号', '产品型号', '客户代码', '客户名称', '分配数量(吨)']
                ws.append(headers)

                # 写入数据
                for _, row in details_df.iterrows():
                    ws.append([
                        row['配方编号'],
                        row['产品型号'],
                        row['客户代码'],
                        row['客户名称'],
                        row['分配数量']
                    ])

                # 设置格式
                for row_idx in range(1, ws.max_row + 1):
                    for col in range(1, 6):
                        cell = ws.cell(row=row_idx, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(size=10)
                        if row_idx > 1 and col == 5:  # 分配数量列
                            cell.number_format = '#,##0.###'

                # 设置列宽
                col_widths = [15, 20, 15, 20, 15]
                for idx, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(idx)].width = width

        # 1. 内蒙未定归属库存（客户为空的库存）
        nm_undetermined = NM_all_storage[NM_all_storage['客户'].str.strip() == ''].copy()
        print("nm_undetermined",nm_undetermined)
        if not nm_undetermined.empty:
            # 添加类型列以便识别
            nm_undetermined['类型'] = '未定归属库存'
            nm_undetermined = nm_undetermined[['产品型号', '库存（吨）']]
            write_unmatched_sheet(wb, nm_undetermined.to_dict('records'),
                                  '内蒙未定归属库存', ['产品型号', '库存（吨）'])

        # 2. 湖北未定归属库存（客户为空的库存）
        hb_undetermined = HB_all_storage[HB_all_storage['客户'].str.strip() == ''].copy()

        print("hb_undetermined", hb_undetermined)
        if not hb_undetermined.empty:
            hb_undetermined['类型'] = '未定归属库存'
            hb_undetermined = hb_undetermined[['产品型号', '库存（吨）']]
            write_unmatched_sheet(wb, hb_undetermined.to_dict('records'),
                                  '湖北未定归属库存', ['产品型号', '库存（吨）'])



        # 写入所有未匹配数据到新sheet
        write_unmatched_sheet(wb, nm_unmatched, '内蒙未在销售计划库存', ['客户', '产品型号', '库存'])
        write_unmatched_sheet(wb, hb_unmatched, '湖北未在销售计划库存', ['客户', '产品型号', '库存'])
        write_unmatched_sheet(wb, nm_send_unmatched, '内蒙需要确认发货', ['客户', '产品型号', '内蒙发货数'])
        write_unmatched_sheet(wb, hb_send_unmatched, '湖北需要确认发货', ['客户', '产品型号', '湖北发货数'])
        write_unmatched_sheet(wb, nm_rate_unmatched, '内蒙配比单未处理', ['客户', '产品型号', '剩余生产量'])
        write_unmatched_sheet(wb, hb_rate_unmatched, '湖北配比单未处理', ['客户', '产品型号', '剩余生产量'])
        print("nm_allocation_details",nm_allocation_details)
        write_allocation_sheet(wb, nm_allocation_details, '内蒙多客户配比单分配统计')
        write_allocation_sheet(wb, hb_allocation_details, '湖北多客户配比单分配统计')

        # 3. 内蒙特殊库存
        if not NM_return_data.empty:
            print("NM_return_data", NM_return_data)
            # 添加类型列以便识别
            NM_return_data['类型'] = '特殊库存'
            # 只保留需要的列
            nm_special_cols = [col for col in NM_return_data.columns if col != '类型']
            nm_special_cols.insert(0, '类型')  # 将类型列放在第一列
            write_unmatched_sheet(wb, NM_return_data.to_dict('records'),
                                  '内蒙特殊库存', nm_special_cols)

        # 4. 湖北特殊库存
        if not HB_return_data.empty:
            print("HB_return_data2",HB_return_data)
            HB_return_data['类型'] = '特殊库存'
            hb_special_cols = [col for col in HB_return_data.columns if col != '类型']
            hb_special_cols.insert(0, '类型')  # 将类型列放在第一列
            write_unmatched_sheet(wb, HB_return_data.to_dict('records'),
                                  '湖北特殊库存', hb_special_cols)

        # 保存修改
        wb.save(new_file_path)
        print(f"成功更新 {updated_rows} 行数据到: {new_file_path}")
        return new_file_path

    except KeyError as e:
        if "内蒙计划数" in str(e) or "湖北计划数" in str(e):
            raise ValueError("sales_df必须包含'内蒙计划数'和'湖北计划数'列") from e
        else:
            raise
    except Exception as e:
        # 清理生成失败的文件
        if os.path.exists(new_file_path):
            os.remove(new_file_path)
        raise RuntimeError(f"文件更新失败: {str(e)}")





def resplitsalesplan(isCreate, original_file_path):
    if isCreate != 1:
        return
    smart_decimal_format = '#,##0.###'
    # 加载工作簿
    wb = openpyxl.load_workbook(original_file_path)

    # 检查工作表是否存在
    if "2025年销售计划" not in wb.sheetnames:
        print(f"错误：工作簿中不存在名为'2025年销售计划'的工作表")
        return

    # 获取原始工作表
    original_sheet = wb["2025年销售计划"]

    # 添加新列
    last_col = original_sheet.max_column
    original_sheet.cell(row=2, column=last_col + 1, value="内蒙")
    original_sheet.cell(row=2, column=last_col + 2, value="湖北")
    original_sheet.cell(row=2, column=last_col + 3, value="标识")

    # 设置标识列的值（注意Excel行号从1开始）
    flag_rows = [8, 33]
    for row in flag_rows:
        if row <= original_sheet.max_row:
            original_sheet.cell(row=row, column=last_col + 3).value = 1

    # 获取当前月份（6月）
    current_month = datetime.now().month
    month_col = None
    for col in range(6, last_col + 1):  # 从F列开始查找
        cell_value = original_sheet.cell(row=2, column=col).value
        if cell_value and str(current_month) + "月" in str(cell_value):
            month_col = col
            break

    if month_col is None:
        print("错误：未找到当月列（1-12月）")
        return

    # 填充内蒙和湖北列
    for row in range(3, original_sheet.max_row + 1):
        customer_name = original_sheet.cell(row=row, column=3).value
        # 检查是否为合计行
        if customer_name and "合计" in str(customer_name):
            continue
        flag_cell = original_sheet.cell(row=row, column=last_col + 3)
        month_value = original_sheet.cell(row=row, column=month_col).value

        original_sheet.cell(row=row, column=last_col + 3).number_format = smart_decimal_format

        if flag_cell.value == 1:
            # 复制到湖北列
            original_sheet.cell(row=row, column=last_col + 2).value = month_value
            original_sheet.cell(row=row, column=last_col + 2).number_format = smart_decimal_format
        else:
            # 复制到内蒙列
            original_sheet.cell(row=row, column=last_col + 1).value = month_value
            original_sheet.cell(row=row, column=last_col + 1).number_format = smart_decimal_format

    # 复制工作表（深拷贝样式）
    # 创建新工作表并复制所有内容
    def copy_sheet_with_style(source, target_name):
        target = wb.create_sheet(target_name)
        for row in source.iter_rows():
            for cell in row:
                new_cell = target.cell(
                    row=cell.row,
                    column=cell.column,
                    value=cell.value
                )
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = smart_decimal_format
                    new_cell.alignment = copy(cell.alignment)
        # 复制列宽
        for col in range(1, source.max_column + 1):
            col_letter = get_column_letter(col)
            # 确保源列维度存在
            if col_letter in source.column_dimensions:
                source_dim = source.column_dimensions[col_letter]

                # 确保目标列维度存在
                if col_letter not in target.column_dimensions:
                    target.column_dimensions[col_letter] = openpyxl.worksheet.dimensions.ColumnDimension(
                        target, index=col_letter
                    )

                # 复制宽度
                if hasattr(source_dim, 'width') and source_dim.width is not None:
                    target.column_dimensions[col_letter].width = source_dim.width
            #target.column_dimensions[col_letter].width = source.column_dimensions[col_letter].width
        return target

    # 创建工作表副本
    neimeng_sheet = copy_sheet_with_style(original_sheet, "2025年内蒙销售计划")
    hubei_sheet = copy_sheet_with_style(original_sheet, "2025年湖北销售计划")

    # === 隐藏E列与当月列之间的列 ===
    # E列是第5列，当月列是month_col
    if month_col > 6:  # 确保有列需要隐藏
        start_hide_col = 6  # F列开始
        end_hide_col = month_col - 1  # 当月列的前一列

        for sheet in [neimeng_sheet, hubei_sheet]:
            for col in range(start_hide_col, end_hide_col + 1):
                col_letter = get_column_letter(col)
                # 确保列维度对象存在
                if col_letter not in sheet.column_dimensions:
                    # 如果不存在则创建
                    sheet.column_dimensions[col_letter] = openpyxl.worksheet.dimensions.ColumnDimension(
                        sheet, index=col_letter
                    )
                sheet.column_dimensions[col_letter].hidden = True

    # 处理内蒙销售计划
    neimeng_sheet = wb["2025年内蒙销售计划"]

    # 查找最后一个"合计"行
    last_total_row = None
    for row in range(neimeng_sheet.max_row, 2, -1):
        if neimeng_sheet.cell(row=row, column=3).value == "合计":
            last_total_row = row
            break

    if last_total_row is None:
        last_total_row = neimeng_sheet.max_row

    # 从高到低删除标识为1的行（在3到合计行之间）
    flag_rows_to_delete = [row for row in range(3, last_total_row + 1)
                           if neimeng_sheet.cell(row=row, column=last_col + 3).value == 1]

    # 按从大到小排序
    flag_rows_to_delete.sort(reverse=True)

    for row in flag_rows_to_delete:
        neimeng_sheet.delete_rows(row)

    # 处理湖北销售计划
    hubei_sheet = wb["2025年湖北销售计划"]

    # 查找最后一个"合计"行
    last_total_row = None
    for row in range(hubei_sheet.max_row, 2, -1):
        if hubei_sheet.cell(row=row, column=3).value == "合计":
            last_total_row = row
            break

    if last_total_row is None:
        last_total_row = hubei_sheet.max_row

    # 从高到低删除非目标行（3到合计行之间）
    rows_to_delete = []
    for row in range(3, last_total_row + 1):
        flag_value = hubei_sheet.cell(row=row, column=last_col + 3).value
        customer_value = str(hubei_sheet.cell(row=row, column=3).value)

        # 保留条件：标识为1 或 包含"合计"
        if flag_value != 1 and "合计" not in customer_value:
            rows_to_delete.append(row)

    # 按从大到小排序
    rows_to_delete.sort(reverse=True)

    for row in rows_to_delete:
        hubei_sheet.delete_rows(row)

    # 保存修改
    wb.save(original_file_path)

    print("操作成功完成！")
    return 1

# 示例调用
#resplitsalesplan(1, "销售2025年销售计划6.13-生产.xlsx")

def GetSalesPlanInit(filePath):
    #销售给的原始文件,清洗合计行、按成品合并销售月计划，并排序
    # 加载工作簿和工作表
    sheet_name = "2025年销售计划"
    wb = openpyxl.load_workbook(filePath, data_only=False)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 '{sheet_name}' 不存在")
    ws = wb[sheet_name]

    # 创建ExcelCompiler对象来计算公式
    compiler = ExcelCompiler(filename=filePath)

    # 确定有效列数：第二行（行号2）中从A列开始直到第一个空列
    max_col = 0
    for cell in ws[2]:  # 第二行
        if cell.value is None:
            break
        max_col += 1

    # 检查是否包含必要的列（D列和F列）
    # if max_col < 4:
    #     raise ValueError("有效列不足，未包含D列（第4列）")
    if max_col < 6:
        raise ValueError("有效列不足，未包含F列（第6列）")

    # 提取第二行的列名（A列到有效列末尾）
    columns = [cell.value for cell in ws[2][:max_col]]

    data = []
    # 从第三行开始读取数据
    for row in ws.iter_rows(min_row=3,values_only=False):
        # 检查停止条件
        a_val = row[0].value if len(row) > 0 else None
        b_val = row[1].value if len(row) > 1 else None
        c_val = row[2].value if len(row) > 2 else None

        # 判断是否满足停止条件
        stop_condition1 = (a_val in (None, '')) and (b_val in (None, '')) and (c_val in (None, ''))
        stop_condition2 = (a_val in (None, '') and (b_val in (None, '')) and c_val == '代加工客户销售计划')
        if stop_condition1 or stop_condition2:
            break

        # 跳过包含“合计”的行
        if c_val and '合计' in str(c_val):
            continue

        # 处理当前行数据
        row_data = []
        for col_idx in range(max_col):
            cell = row[col_idx] if col_idx < len(row) else None
            if cell.data_type == 'f':  # 如果单元格包含公式
                # 使用pycel计算公式
                formula = cell.value
                # 转换单元格坐标为A1表示法
                cell_address = f'{sheet_name}!{cell.coordinate}'
                try:
                    cell_value = compiler.evaluate(cell_address)
                    #row_data.append(result)
                except Exception as e:
                    print(f"公式计算错误: {formula} 在 {cell_address}, 错误信息: {e}")
                    #row_data.append(None)
                    cell_value = None
            else:
                cell_value = cell.value


            #cell_value = cell.value if cell is not None else None

            # 处理公式值
            # if cell and cell.data_type == 'f':
            #     cell_value = calculate_formula_value(cell)
            # if cell:
            #     cell_value = convert_formula(cell)

            # 修改D列（索引3）的值
            if col_idx == 3 and cell_value == 'SN-LTF':
                cell_value = 'SN-P2C-1'

            row_data.append(cell_value)

        data.append(row_data)

    # 转换为DataFrame
    df = pd.DataFrame(data, columns=columns)
    original_columns = df.columns.tolist()  # 获取原始列顺序
    month_columns = [col for col in original_columns if re.match(r'^\d+月$', col)]
    if df.empty:
        return pd.DataFrame()

    #定义D列的自定义排序规则
    custom_order3 = [
        'SN-P1C', 'SN-P2C', 'SN-P1L','SN-P1G','SN-P2C-1', 'SN-4D-1',
        'SN-H01', 'SN-H04', 'MAG-4CL','SN-P2H','SN-K1-3T', 'SN-09G',
        'MAG-106','SN-P2', 'SN-BP12T','SN-BP12','SN-DA-3', 'SN-C1M',
        'SN-P1','SN-K1M','SN-P2C-2','SN-P2C-GX', 'SN-DA-3A','SN-DA-3L',
         'SN-8B-BL','SN-BN1','SN-P1L-A',  'SN-C1M-A'
    ] #SN-P1L,SN-H04 是配比有，而销售计划无的型号，'SN-DA-3', 'SN-C1M','MAG-507', 'MAG-P3',','SN-SF3','MAG-100','MAG-10B','SN-P2F', 'MAG-P2A', 'SN-P7'
    custom_order3 = [
        'SN-P1C', 'SN-P2C', 'SN-P1G', 'SN-P2C-1', 'SN-4D-1',
        'SN-H01', 'MAG-4CL', 'SN-P2H', 'SN-K1-3T', 'SN-09G',
        'MAG-106', 'SN-P2', 'SN-BP12T', 'SN-BP12',
        'SN-P1', 'SN-K1M', 'SN-P2C-2', 'SN-P2C-GX', 'SN-DA-3A', 'SN-DA-3L',
        'SN-8B-BL', 'SN-BN1', 'SN-P1L-A', 'SN-C1M-A'
    ]

    custom_order = [
        'SN-P1C', 'SN-P2C', 'SN-P1L', 'SN-P1G', 'SN-P2C-1', 'SN-4D-1',
        'SN-H01', 'SN-H04', 'MAG-4CL', 'SN-P2H', 'MAG-106A', 'SN-10BL',
        'MAG-09', 'SN-09G', 'MAG-106', 'SN-P2', 'SN-BP12T', 'SN-BP12',
        'SN-BN1', 'SN-K1-3T','SN-C1M-A','SN-P1', 'SN-K1M', 'SN-P2C-2',
        'SN-P2C-GX'
    ]


    # 将D列转换为分类类型，并指定顺序 , 'SN-DK' 'SN-DA','SN-DA-4'
    df[columns[3]] = pd.Categorical(
        df[columns[3]],
        categories=custom_order,
        ordered=True
    )



    # 确定分组列（D列）和需要聚合的列（F列及之后）
    group_col = columns[3]
    agg_columns = columns[5:]  # F列是第6列，索引5

    # 确保聚合列为数值类型
    for col in agg_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # 确保目标列名存在
    target_column = '每月目标安全库存'
    if target_column not in columns:
        raise ValueError(f"列 '{target_column}' 不存在于数据中")

    # 确定分组列（D列）和需要聚合的列（包括目标列）
    group_col = columns[3]
    # 显式指定需要聚合的月份列和目标列
    #month_columns = ['1月', '2月', '3月', '4月', '5月', '6月', '7月','8月']



    agg_columns = month_columns + [target_column]

    # 确保列名存在于DataFrame中
    missing_cols = [col for col in agg_columns if col not in df.columns]
    if missing_cols:
        raise ValueError(f"缺失列: {missing_cols}")

    # 转换为数值类型，处理空值
    for col in agg_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # 按D列分组并聚合求和
    #grouped_df = df.groupby(group_col, as_index=False)[agg_columns].sum()
    grouped_df = df.groupby(group_col, as_index=False, observed=False)[agg_columns].sum()

    # 按自定义顺序排序
    grouped_df = grouped_df.sort_values(by=group_col)

    # 重置索引（可选）
    grouped_df.reset_index(drop=True, inplace=True)

    return grouped_df


df = GetSalesPlanInit("销售2025年销售计划6.13-生产.xlsx")
print("df",df[['产品型号','6月']])


def process_workbook_results(marrequirfilename, factory, sales_df, radio_df, dailyfilename=None,logic=None):
    # 加载主需求工作簿
    wb = load_workbook(marrequirfilename)
    sheets = wb.sheetnames

    # 初始化数据容器
    YCL_data = pd.DataFrame(columns=["operation", "factory", "months", "totalnum"])
    FH_data = pd.DataFrame(columns=["operation", "factory", "months", "totalnum"])
    DWTH_data = pd.DataFrame(columns=["operation", "factory", "months", "totalnum"])  # 内蒙低温炭化
    #DWTH_data_hb = pd.DataFrame(columns=["operation", "factory", "months", "totalnum"])  # 湖北低温炭化
    SMH_data = pd.DataFrame(columns=["operation", "factory", "months", "totalnum"])
    GWTH_data = pd.DataFrame(columns=["operation", "factory", "months", "totalnum"])

    special_col_rules = {
        'P': ('O', 'P'),  # 2-1T#: O列是2-1#, P列是2-1T#
        'Q': ('R', 'Q'),  # 2A-1TB2#: Q列是2A-1TB2#, R列是2A-2#
        'V': ('U', 'V'),  # 石S1-1T#: U列是石S1-1#, V列是石S1-1T#
        'Y': ('X', 'Y'),  # 针S1-1T#: X列是针S1-1#, Y列是针S1-1T#
        'AA': ('AA', 'AA'),  # S1-3T#: Z列是S1-2#, AA列是S1-3T#
        'AC': ('AB', 'AC')  # S2-2T#: AB列是高S2-2#, AC列是S2-2T#
    }

    SMH_radio = pd.DataFrame(columns=common_columns)

    # 高温炭化特殊列
    high_temp_special_cols = ['P', 'Q', 'V', 'Y', 'AA', 'AC']

    # ----------- 核心工具函数 -----------
    def get_merged_value(sheet, row, col):
        """处理合并单元格，返回左上角的值"""
        cell = sheet.cell(row=row, column=col)
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return sheet.cell(merged_range.min_row, merged_range.min_col).value
        return cell.value

    def safe_float(value):
        """增强数值转换逻辑"""
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace(' ', '')
                if '%' in value:
                    return float(value.replace('%', '')) / 100
            return float(value) if value not in ("/", None, "", " ") else 0.0
        except (ValueError, TypeError) as e:
            print(f"数值转换异常：{str(e)} → 原始值：'{value}'")
            return 0.0

    def offset_month(base_month, offset):
        """处理月份偏移，支持跨年"""
        month_num = int(base_month.replace("月", ""))
        new_month = (month_num + offset - 1) % 12 + 1
        return f"{new_month}月"

    # ----------- 主处理逻辑 -----------
    valid_months = [f"{i}月" for i in range(5, 13)]  # 5月到12月

    SMH_radio = pd.DataFrame(columns=common_columns)
    DWTH_radio = pd.DataFrame(columns=common_columns)
    GWTH_radio = pd.DataFrame(columns=common_columns)

    sheet_idx = 0

    preMonth = "1月"
    for sheet_name in sheets:
        if "原材料需求倒算" not in sheet_name:
            continue

        current_month = sheet_name.replace("原材料需求倒算", "")
        if current_month not in valid_months:
            continue

        sheet = wb[sheet_name]
        if sheet_idx == 0:
            preMonth = current_month

        # === 构建SMH_radio数据 ===
        DWTH_radio_values = []
        for col_idx in range(13, 35):  # Excel列号从1开始，M=13，AH=34
            raw_value = get_merged_value(sheet, 56, col_idx)
            clean_value = safe_float(raw_value)
            DWTH_radio_values.append(clean_value)

            # 将读取的值按common_columns顺序插入DataFrame
            # 注意：需确保common_columns长度与列数一致（22列）
        SMH_radio_values = []
        for col_idx in range(13, 35):  # Excel列号从1开始，M=13，AH=34
            raw_value = get_merged_value(sheet, 48, col_idx)
            clean_value = safe_float(raw_value)
            SMH_radio_values.append(clean_value)

        GWTH_radio_values = []
        for col_idx in range(13, 35):  # Excel列号从1开始，M=13，AH=34
            raw_value = get_merged_value(sheet, 40, col_idx)
            clean_value = safe_float(raw_value)
            GWTH_radio_values.append(clean_value)

        print("GWTH_radio_values",GWTH_radio_values)

        if len(DWTH_radio_values) == len(common_columns):
            DWTH_radio = pd.concat([DWTH_radio, pd.DataFrame([DWTH_radio_values], columns=common_columns)],ignore_index=True)
        else:
            print(f"警告：Sheet {sheet_name} 的第56行列数不匹配，已跳过")

        if len(SMH_radio_values) == len(common_columns):
            SMH_radio = pd.concat([SMH_radio, pd.DataFrame([SMH_radio_values], columns=common_columns)],ignore_index=True)
        else:
            print(f"警告：Sheet {sheet_name} 的第48行列数不匹配，已跳过")

        if len(GWTH_radio_values) == len(common_columns):
            GWTH_radio = pd.concat([GWTH_radio, pd.DataFrame([GWTH_radio_values], columns=common_columns)],ignore_index=True)
        else:
            print(f"警告：Sheet {sheet_name} 的第40行列数不匹配，已跳过")




        # target_cols_75 = ['M', 'N', 'O', 'Q', 'S', 'T', 'U', 'X', 'W', 'Z', 'AA', 'AB', 'AD', 'AE', 'AF', 'AG',
        #                   'AH']
        target_cols_71 = ['M', 'N', 'O','P', 'Q', 'R','S', 'T', 'U','V',  'W', 'X','Y','Z', 'AA', 'AB','AC', 'AD', 'AE', 'AF', 'AG',
                          'AH']
        # === 各工序数据计算 ===
        def calculate_product(row_numbers, operation=None):
            """通用乘积计算函数，支持特殊列处理"""
            total = 0.0
            multiply_count = 0  # 记录有效正数相乘的次数

            for col_letter in target_cols_71:
                product = 1.0
                col_idx = column_index_from_string(col_letter)

                nindex = 0
                valid = True  # 标记当前列是否有效
                for row in row_numbers:
                    # 特殊处理高温炭化工序的特殊列
                    if operation == "高温炭化":
                        if col_letter in high_temp_special_cols:
                            # 获取相关列
                            main_col, dep_col = special_col_rules[col_letter]
                            main_col_idx = column_index_from_string(main_col)
                            dep_col_idx = column_index_from_string(dep_col)

                            # 读取主列和依赖列的值
                            main_val = safe_float(get_merged_value(sheet, row, main_col_idx))
                            dep_val = safe_float(get_merged_value(sheet, row, dep_col_idx))

                            if col_letter == "P" and (row == 75 or row == 68):
                                valid = False
                                break
                            if col_letter == "R" and (row == 75 or row == 68):
                                valid = False
                                break
                            if col_letter == "V" and (row == 75 or row == 68):
                                valid = False
                                break
                            if col_letter == "Y" and (row == 75 or row == 68):
                                valid = False
                                break
                            if col_letter == "AC" and (row == 75 or row == 68):
                                valid = False
                                break

                            # 如果是第一个sheet，高炭不需要考虑 + 72行的逻辑
                            if sheet_idx == 0:
                                # 读取第72行的值
                                main_val72 = 0 #safe_float(get_merged_value(sheet, 72, main_col_idx))
                                dep_val72 = 0 #safe_float(get_merged_value(sheet, 72, dep_col_idx))

                                # 如果值大于0，则相加
                                if main_val > 0 and main_val72 > 0:
                                    main_val += main_val72
                                if dep_val > 0 and dep_val72 > 0:
                                    dep_val += dep_val72

                                if col_letter=="P" and row == 56:
                                    continue
                                if col_letter == "V" and row == 62:
                                    continue
                                if col_letter == "Y" and row == 62:
                                    continue
                                if col_letter == "V" and row == 62:
                                    continue
                                if col_letter == "AA" and row == 62:
                                    continue

                                # 使用相加后的值
                                value =  dep_val #main_val +
                                #print("gaotan-idx0",main_val,dep_val)
                            else:

                                # 对于后续的sheet，使用71行的值,参考74行的值（结余值），如果结余值74大于 71值，则71值0，否则71值-74值
                                main_val74 = safe_float(get_merged_value(sheet, 74, main_col_idx))
                                dep_val74 = safe_float(get_merged_value(sheet, 74, dep_col_idx))

                                # 计算差值（只取正值）
                                main_val = max(0, main_val - main_val74)
                                dep_val = max(0, dep_val - dep_val74)

                                value =  dep_val #main_val +

                                if col_letter=="P" and row == 56:
                                    continue
                                if col_letter=="V" and row == 62:
                                    continue
                                if col_letter == "Y" and row == 62:
                                    continue
                                if col_letter == "AA" and row == 62:
                                    continue

                            if nindex == 0 and value <= 0:
                                valid = False
                                break
                            if nindex > 0 and (value <= 0 or value >= 1):
                                valid = False
                                break

                            product *= value
                            multiply_count += 1
                            #print("calculate_product-gaotan", col_letter, col_idx,sheet_idx, row,dep_val,main_val,value,product)
                        else:
                            continue
                    else:
                        # 普通列的处理
                        raw_value = get_merged_value(sheet, row, col_idx)
                        clean_value = safe_float(raw_value) #获得第71行col_idx列数据

                        if col_letter == "P" and (row == 75 or row == 68):
                            valid = False
                            break
                        if col_letter == "R" and (row == 75 or row == 68):
                            valid = False
                            break
                        if col_letter == "V" and (row == 75 or row == 68):
                            valid = False
                            break
                        if col_letter == "Y" and (row == 75 or row == 68):
                            valid = False
                            break
                        if col_letter == "AC" and (row == 75 or row == 68):
                            valid = False
                            break
                        print("row-feigaotan1",row,col_letter,raw_value,clean_value)

                        # 如果是第一个sheet且是71行，检查是否有72行需要相加
                        if sheet_idx == 0:
                            #raw_value72 = get_merged_value(sheet, 72, col_idx)
                            #clean_value72 = safe_float(raw_value72)

                            if clean_value > 0:
                                clean_value += 0
                        else:
                            # 对于后续的sheet，使用71行的值减去74行的值
                            main_val74 = safe_float(get_merged_value(sheet, 74, col_idx))
                            if clean_value > 0 and main_val74 >0 and clean_value < main_val74:
                                clean_value = 0
                            elif clean_value > 0 and main_val74 >0 and clean_value >= main_val74:
                                clean_value = clean_value - main_val74
                            #对于后续的行

                        if col_letter == "M" and (row == 40 or row ==56 or row==62):
                            continue

                        if col_letter == "N" and (row == 40 or row ==56 or row==62):
                            continue
                        if col_letter == "O" and (row == 40 or row ==56):
                            continue
                        if col_letter == "P" and (row == 40 or row ==56):
                            continue

                        if nindex == 0 and clean_value <= 0:
                            valid = False
                            break
                        if nindex > 0 and (clean_value <= 0 or clean_value >= 1):
                            valid = False
                            break

                        product *= clean_value
                        multiply_count += 1
                        print("row-feigaotan2", row, col_letter, raw_value, clean_value,product)
                        append_data_using_open(
                            f"Operation: {operation}, Column: {col_letter}, Multiply Count: {multiply_count}, clean_value: {clean_value}, product: {product}\n")
                        #if nindex ==0:
                            #print("calculate_product-feigaotan", col_letter, col_idx,  clean_value, product)
                    nindex += 1

                # if product > 0:
                #     total += product

                if valid and product > 0:
                    total += product

                    print(f"Operation: {operation}, Column: {col_letter}, Multiply Count: {multiply_count}, Total: {total}")
                    append_data_using_open(f"Operation: {operation}, Column: {col_letter}, Multiply Count: {multiply_count}, Total: {total}\n")

            #print("row-feigaotan3", total)
            print(f"Operation: {operation}, Final Multiply Count: {multiply_count}, Final Total: {total}")
            append_data_using_open(f"Operation: {operation}, Final Multiply Count: {multiply_count}, Final Total: {total:.1f}\n")
            return int(total)

        # 计算各工序总量
        print("calculate_product-预处理",)
        YCL_total = calculate_product([75, 68],"预处理")
        print("calculate_product-复合",)
        FH_total = calculate_product([75, 68, 62],"复合")
        print("calculate_product-低温炭化", )
        DWTH_total = calculate_product([75, 68, 62, 56],"低温炭化")
        print("calculate_product-石墨化", )
        SMH_total = calculate_product([75, 68, 62, 56, 48],"石墨化")
        print("calculate_product-高温炭化", )
        GWTH_total = calculate_product([75, 68, 62, 56, 48,40], operation="高温炭化")

        print("new_data-currentmonth",current_month)
        # 存储基础数据
        new_data = {
            "预处理": (YCL_data, YCL_total, current_month),
            "复合造粒": (FH_data, FH_total, current_month),
            "低温炭化": (DWTH_data, DWTH_total, offset_month(current_month, 1)),
            "石墨化":  (SMH_data, SMH_total, offset_month(current_month, 1)),
            "高温炭化": (GWTH_data, GWTH_total, offset_month(current_month, 1))
        }

        for op, (df, total, month) in new_data.items():
            df.loc[len(df)] = {
                "operation": op,
                "factory": factory,
                "months": month,
                "totalnum": total
            }
            #if op == "低温炭化" and month=="6月"
            print("get-value",op,factory,month,total)

        sheet_idx += 1

        # 分工厂存储低温炭化数据
        dwth_month = offset_month(current_month, 1)


    # === 特殊日报表处理 ===
    print("factory",factory,dailyfilename)
    if factory in ["内蒙", "湖北","所有"] and dailyfilename:
        try:
            daily_wb = load_workbook(dailyfilename)
            sheet_map = {
                "内蒙": "库存汇总",
                "湖北": "库存汇总",
                "所有": "库存汇总"
            }
            daily_sheet = daily_wb[sheet_map[factory]]

            current_models = sales_df["产品型号"].tolist()  # 获取当前处理的产品型号列表
            filtered_radio_df = radio_df[radio_df["产品型号"].isin(current_models)]

            # 移除'产品型号'列并验证列数
            filtered_radio_df = filtered_radio_df.drop(columns=['产品型号'])
            if filtered_radio_df.shape[1] != len(common_columns):
                raise ValueError(
                    f"radio_df列数({filtered_radio_df.shape[1]})与common_columns({len(common_columns)})不匹配")

            # 转换为NumPy数组
            radio_matrix = filtered_radio_df.values  # 形状应为 (n, 22)

            # === 通用处理函数 ===
            def process_daily_data(row_num, radio_matrix, target_radio_df, target_col_letters=None):
                """处理日报表指定行数据并返回计算结果"""
                if target_col_letters is None:
                    target_col_letters = [
                        'B', 'C', 'D', 'E', 'F', 'H', 'I', 'J', 'K', 'L',
                        'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'AA', 'AB'
                    ]

                daily_values = []
                for col_letter in target_col_letters:
                    col_idx = column_index_from_string(col_letter)
                    raw_value = get_merged_value(daily_sheet, row_num, col_idx)
                    clean_value = safe_float(raw_value)
                    daily_values.append(clean_value)
                print("daily_values",daily_values)

                if len(daily_values) == len(common_columns):
                    daily_array = np.array(daily_values).reshape(1, -1)
                    target_radio = target_radio_df.iloc[-1].values.reshape(1, -1)
                    matrix_result = daily_array  * target_radio  #* radio_matrix
                    print("zailu",daily_array * target_radio)
                    print("zailutotal",int(np.sum(matrix_result[matrix_result > 0])))
                    return int(np.sum(matrix_result[matrix_result > 0]))
                else:
                    print(f"日报表列数不匹配，已跳过计算")
                    return 0

            # === 处理SMH_data（原逻辑） ===
            smh_total = process_daily_data(20, radio_matrix, SMH_radio)
            SMH_data.loc[len(SMH_data)] = {
                "operation": "石墨化",
                "factory": factory,
                "months": preMonth,
                "totalnum": smh_total
            }
            print("SMH_data-smh_total-radio_matrix", radio_matrix)
            print("SMH_data-smh_total-SMH_radio", SMH_radio)
            print("SMH_data-smh_total",smh_total)
            print("SMH_data",SMH_data)
            # === 新增DWTH_data处理（第16行） ===
            dwth_total = process_daily_data(16, radio_matrix, DWTH_radio)
            DWTH_data.loc[len(DWTH_data)] = {
                "operation": "低温炭化",
                "factory": factory,
                "months": preMonth,
                "totalnum": dwth_total
            }

            # === 新增GWTH_data处理（第28行，修改F列为G列） ===
            gwth_col_letters = [
                'B', 'C', 'D', 'E', 'G', 'H', 'I', 'J', 'K', 'L',  # F改为G
                'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'AA', 'AB'
            ]
            gwth_total = process_daily_data(28, radio_matrix, GWTH_radio, gwth_col_letters)
            GWTH_data.loc[len(GWTH_data)] = {
                "operation": "高温炭化",
                "factory": factory,
                "months": preMonth,
                "totalnum": gwth_total
            }

            # 数据排序处理
            for df in [SMH_data, DWTH_data, GWTH_data]:
                df['months'] = df['months'].astype(str)
                month_order = {f"{i}月": i for i in range(1, 13)}
                df['month_num'] = df['months'].map(month_order)
                df.sort_values('month_num', inplace=True)
                df.drop('month_num', axis=1, inplace=True)
                df.reset_index(drop=True, inplace=True)



        except Exception as e:
            print(f"日报表处理失败：{str(e)}")

    # ----------- 数据补全逻辑 -----------
    def complete_months(df, offset=False):
        """自动补全至12月数据"""
        if df.empty:
            return df

        last_month = df.iloc[-1]["months"]
        last_num = int(last_month.replace("月", ""))

        new_rows = []
        for m in range(last_num + 1, 13):
            new_row = df.iloc[-1].copy()
            new_month = offset_month(last_month, m - last_num) if offset else f"{m}月"
            new_row["months"] = new_month
            new_rows.append(new_row)

        return pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True) if new_rows else df

    # 返回对应工厂数据集
    DWTH_result = DWTH_data # if factory == "内蒙" else DWTH_data_hb
    return (
        complete_months(YCL_data),
        complete_months(FH_data),
        complete_months(DWTH_result, offset=True),
        complete_months(SMH_data, offset=True),
        complete_months(GWTH_data,offset=True)
    )
